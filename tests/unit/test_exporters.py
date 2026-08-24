"""Tests for report generation, including untrusted-data handling."""

import csv
import json

import pytest
from openpyxl import load_workbook

from exporters import (
    CSVExporter,
    ExcelExporter,
    HTMLExporter,
    JSONExporter,
    format_threat_intel,
    join_values,
)

# A permutation whose WHOIS record and page title are controlled by the
# squatter being investigated. Every field here is hostile by construction.
HOSTILE_PERM = {
    'domain': 'evil<img src=x onerror=alert(1)>.com',
    'fuzzer': 'homoglyph',
    'risk_score': 90,
    'created_days_ago': 2,
    'is_recent': True,
    'dns_a': ['203.0.113.1'],
    'whois_registrant': '<script>alert("xss")</script>',
    'whois_org': '=cmd|" /C calc"!A0',
    'whois_created': ['2026-01-01'],
    'threat_intel': {
        'urlscan': {
            'malicious': True,
            'score': 100,
            'report_url': 'javascript:alert(document.cookie)',
        },
        'http_probe': {'https_active': True, 'https_status': 200,
                       'title': '</td></tr><script>alert(2)</script>'},
    },
}

HOSTILE_RESULTS = [{
    'original_domain': 'example.com',
    'scan_date': '2026-01-15',
    'total_permutations': 1,
    'registered_count': 1,
    'filtered_count': 1,
    'permutations': [HOSTILE_PERM],
}]


class TestHtmlExporterEscaping:
    @pytest.fixture
    def report(self, config, tmp_path):
        path = HTMLExporter(config).export(HOSTILE_RESULTS, tmp_path)
        return path.read_text(encoding='utf-8')

    def test_script_tags_are_escaped(self, report):
        assert '<script>alert("xss")</script>' not in report
        assert '&lt;script&gt;' in report

    def test_domain_markup_is_escaped(self, report):
        assert '<img src=x onerror=alert(1)>' not in report

    def test_javascript_urls_never_become_links(self, report):
        assert 'javascript:' not in report

    def test_external_links_are_isolated(self, config, tmp_path):
        safe = json.loads(json.dumps(HOSTILE_RESULTS))
        safe[0]['permutations'][0]['threat_intel']['urlscan']['report_url'] = \
            'https://urlscan.io/result/abc/'
        report = HTMLExporter(config).export(safe, tmp_path).read_text()
        assert 'rel="noopener noreferrer"' in report

    def test_report_is_still_well_formed(self, report):
        assert report.strip().startswith('<!DOCTYPE html>')
        assert report.strip().endswith('</html>')


class TestCsvExporterInjection:
    @pytest.fixture
    def rows(self, config, tmp_path):
        path = CSVExporter(config).export(HOSTILE_RESULTS, tmp_path)
        with path.open(newline='', encoding='utf-8') as f:
            return list(csv.reader(f))

    def test_formula_cells_are_neutralised(self, rows):
        header, data = rows[0], rows[1]
        org = data[header.index('Organization')]
        assert org.startswith("'=")

    def test_all_expected_columns_present(self, rows):
        for column in ('Risk Score', 'Age (days)', 'URLScan Report', 'Recent'):
            assert column in rows[0]

    def test_values_land_in_the_right_columns(self, rows):
        header, data = rows[0], rows[1]
        assert data[header.index('Risk Score')] == '90'
        assert data[header.index('Age (days)')] == '2'
        assert data[header.index('Recent')] == 'Yes'


class TestExcelExporter:
    @pytest.fixture
    def workbook(self, config, tmp_path):
        return load_workbook(ExcelExporter(config).export(HOSTILE_RESULTS, tmp_path))

    def test_expected_sheets(self, workbook):
        assert workbook.sheetnames == ['Summary', 'Details', 'Statistics']

    def test_formula_cells_are_neutralised(self, workbook):
        values = [c.value for row in workbook['Details'].iter_rows() for c in row]
        assert any(isinstance(v, str) and v.startswith("'=cmd") for v in values)
        assert not any(isinstance(v, str) and v.startswith('=cmd') for v in values)

    def test_column_widths_are_set(self, workbook):
        """Numeric columns used to be skipped by the width calculation."""
        widths = workbook['Details'].column_dimensions
        assert widths['A'].width > 0


class TestJsonExporter:
    def test_round_trips(self, config, tmp_path, sample_results):
        path = JSONExporter(config).export(sample_results, tmp_path)
        data = json.loads(path.read_text(encoding='utf-8'))
        assert data['total_domains'] == 1
        assert data['results'][0]['original_domain'] == 'example.com'
        assert data['version'] != '1.0'  # now tracks the real package version


class TestFormatThreatIntel:
    def test_empty_intel(self):
        assert format_threat_intel({}) == {
            'urlscan': '', 'urlscan_url': None, 'ct': '', 'http': '', 'tls': ''
        }

    def test_urlscan_error_statuses(self):
        out = format_threat_intel({'threat_intel': {'urlscan': {'status': 'rate_limited'}}})
        assert out['urlscan'] == 'Rate Limited'

    def test_unsafe_report_url_is_dropped(self):
        out = format_threat_intel({'threat_intel': {
            'urlscan': {'malicious': True, 'score': 50, 'report_url': 'javascript:x'}}})
        assert out['urlscan_url'] is None
        assert out['urlscan'] == 'Malicious (50)'

    def test_http_prefers_https(self):
        out = format_threat_intel({'threat_intel': {'http_probe': {
            'https_active': True, 'https_status': 200,
            'http_active': True, 'http_status': 301}}})
        assert out['http'] == 'HTTPS: 200'

    @pytest.mark.parametrize('verified,expected', [
        (True, 'Valid'),
        (False, 'Invalid/self-signed'),
        (None, ''),
    ])
    def test_reports_certificate_validity(self, verified, expected):
        out = format_threat_intel({'threat_intel': {'http_probe': {
            'https_active': True, 'https_status': 200, 'tls_verified': verified}}})
        assert out['tls'] == expected

    def test_ct_certificate_count(self):
        out = format_threat_intel({'threat_intel': {
            'certificate_transparency': {'certificates_found': 3}}})
        assert out['ct'] == '3 cert(s)'


class TestJoinValues:
    @pytest.mark.parametrize('value,expected', [
        (None, ''), ([], ''), ('a', 'a'),
        (['a', 'b'], 'a, b'), (['a', None, 'b'], 'a, b'), (42, '42'),
    ])
    def test_join(self, value, expected):
        assert join_values(value) == expected
