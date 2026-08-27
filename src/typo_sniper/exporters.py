"""
Export modules for Typo Sniper results.

Supports multiple output formats: Excel, JSON, CSV, and HTML.
"""

# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 ChiefGyk3D
# Typo Sniper is dual-licensed; see COMMERCIAL.md for commercial terms.

import csv
import html
import json
import logging
from abc import ABC, abstractmethod
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import Config
from .utils import safe_url, sanitize_spreadsheet_value
from .version import __version__


def format_threat_intel(perm: dict[str, Any]) -> dict[str, Any]:
    """
    Render a permutation's threat intelligence into display-ready strings.

    Every exporter needs the same URLScan/CT/HTTP summaries. They used to be
    built by three near-identical inline blocks, which is how the formats
    drifted apart. This is the single source of truth for all of them.

    Args:
        perm: Permutation dictionary

    Returns:
        Dictionary with 'urlscan', 'urlscan_url', 'ct', and 'http' keys
    """
    summary = {'urlscan': '', 'urlscan_url': None, 'ct': '', 'http': '', 'tls': '',
               'mail': '', 'page': ''}

    # Mail capability lives beside threat_intel, not inside it
    mail = perm.get('mail_intel') or {}
    if mail:
        posture = mail.get('posture')
        labels = {
            'none': '',
            'receive-only': 'Receive only (MX)',
            'partial': 'Partial',
            'provisioned': 'SEND-CAPABLE',
            'hardened': 'SEND-CAPABLE (DMARC enforced)',
            'unknown': 'Lookup failed',
        }
        summary['mail'] = labels.get(posture, posture or '')

    threat_intel = perm.get('threat_intel') or {}
    if not threat_intel:
        return summary

    # --- URLScan ---
    us_data = threat_intel.get('urlscan')
    if us_data:
        status = us_data.get('status')
        if status:
            error = us_data.get('error', '')
            labels = {
                'rate_limited': 'Rate Limited',
                'timeout': 'Scan Timeout',
                'submission_failed': f'Scan Failed: {error}' if error else 'Scan Failed',
                'error': f'Error: {error}' if error else 'Error',
            }
            summary['urlscan'] = labels.get(status, f'Error: {status}')
        else:
            malicious = us_data.get('malicious', False)
            score = us_data.get('score', 0)
            summary['urlscan_url'] = safe_url(us_data.get('report_url'))
            verdict = 'Malicious' if (malicious or score > 0) else 'Clean'
            summary['urlscan'] = f'{verdict} ({score})'
            if not summary['urlscan_url'] and verdict == 'Clean':
                summary['urlscan'] = 'No Scan Available'

    # --- Certificate Transparency ---
    ct_data = threat_intel.get('certificate_transparency')
    if ct_data:
        cert_count = ct_data.get('certificates_found', 0)
        if cert_count > 0:
            summary['ct'] = f'{cert_count} cert(s)'
        else:
            summary['ct'] = ct_data.get('status', '') or '0'

    # --- What the page collects ---
    # Derived by our own parser from the fetched markup, so unlike the page
    # title this line is not attacker-authored text.
    page_data = (threat_intel.get('http_probe') or {}).get('page')
    if page_data and page_data.get('parse_ok'):
        from .page_analysis import describe

        summary['page'] = describe(page_data)

    # --- HTTP probe ---
    http_data = threat_intel.get('http_probe')
    if http_data:
        https_code = http_data.get('https_status')
        http_code = http_data.get('http_status')
        if http_data.get('https_active') and https_code:
            summary['http'] = f'HTTPS: {https_code}'
        elif http_data.get('http_active') and http_code:
            summary['http'] = f'HTTP: {http_code}'
        elif http_data.get('https_active'):
            # Answered TLS but the certificate was rejected, so no request
            # was completed and there is no status code to show
            summary['http'] = 'HTTPS: cert rejected'
        else:
            summary['http'] = 'Inactive'

        tls_verified = http_data.get('tls_verified')
        if tls_verified is True:
            summary['tls'] = 'Valid'
        elif tls_verified is False:
            summary['tls'] = 'Invalid/self-signed'

    return summary


def join_values(values: Any) -> str:
    """Join a WHOIS/DNS list field into a single comma-separated string."""
    if not values:
        return ''
    if isinstance(values, str):
        return values
    if isinstance(values, (list, tuple, set)):
        return ', '.join(str(v) for v in values if v)
    return str(values)


class BaseExporter(ABC):
    """Base class for all exporters."""

    def __init__(self, config: Config):
        """
        Initialize exporter.

        Args:
            config: Configuration object
        """
        self.config = config
        self.logger = logging.getLogger(self.__class__.__name__)

    @abstractmethod
    def export(self, results: list[dict[str, Any]], output_dir: Path) -> Path:
        """
        Export results to file.

        Args:
            results: List of scan result dictionaries
            output_dir: Directory to save output file

        Returns:
            Path to the created file
        """
        pass

    def _generate_filename(self, output_dir: Path, extension: str) -> Path:
        """
        Generate output filename with timestamp.

        Args:
            output_dir: Output directory
            extension: File extension (without dot)

        Returns:
            Full path to output file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return output_dir / f"typo_sniper_results_{timestamp}.{extension}"


class ExcelExporter(BaseExporter):
    """Export results to Excel format with rich formatting."""

    def export(self, results: list[dict[str, Any]], output_dir: Path) -> Path:
        """Export results to Excel file."""
        output_file = self._generate_filename(output_dir, 'xlsx')
        
        wb = Workbook()
        
        # Create summary sheet
        self._create_summary_sheet(wb, results)
        
        # Create detailed results sheet
        self._create_details_sheet(wb, results)
        
        # Create statistics sheet
        self._create_statistics_sheet(wb, results)
        
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(output_file)
        self.logger.info(f"Exported Excel file: {output_file}")
        return output_file

    def _create_summary_sheet(self, wb: Workbook, results: list[dict[str, Any]]) -> None:
        """Create summary sheet."""
        ws = wb.active
        ws.title = "Summary"
        
        # Headers
        headers = [
            "Scan Date", "Original Domain", "Total Permutations",
            "Registered", "Filtered", "Recent", "WHOIS OK", "WHOIS Failed"
        ]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for result in results:
            recent_count = len([p for p in result['permutations'] if p.get('is_recent', False)])
            ws.append([
                result['scan_date'],
                result['original_domain'],
                result['total_permutations'],
                result['registered_count'],
                result['filtered_count'],
                recent_count if recent_count > 0 else '',
                result.get('whois_succeeded', ''),
                result.get('whois_failed', ''),
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_details_sheet(self, wb: Workbook, results: list[dict[str, Any]]) -> None:
        """Create detailed results sheet."""
        ws = wb.create_sheet("Details")
        
        # Headers
        headers = [
            "Scan Date", "Original Domain", "Permutation", "Fuzzer Type",
            "Risk Score", "Age (days)", "Mail", "Page", "URLScan Status", "CT Logs",
            "HTTP Status", "TLS",
            "Created Date", "Updated Date", "Expires Date",
            "Registrant", "Organization", "Registrar",
            "Email", "Country", "Status",
            "Name Servers", "IP Address", "Mail Server"
        ]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for result in results:
            scan_date = result['scan_date']
            original = result['original_domain']
            
            # Add spacing row for each domain
            if ws.max_row > 1:
                ws.append([''] * len(headers))
            
            # Add original domain header
            domain_row = [scan_date, original, '*ORIGINAL*'] + [''] * (len(headers) - 3)
            ws.append(domain_row)
            domain_font = Font(bold=True, color="0066CC", size=11)
            for cell in ws[ws.max_row]:
                cell.font = domain_font
            
            # Add permutations
            for perm in result['permutations']:
                intel = format_threat_intel(perm)
                risk_score = perm.get('risk_score', '')

                urlscan_status = intel['urlscan']
                if intel['urlscan_url']:
                    urlscan_status = f"{urlscan_status} - {intel['urlscan_url']}"

                row = [
                    scan_date,
                    original,
                    perm['domain'],
                    perm.get('fuzzer', ''),
                    risk_score,
                    perm.get('created_days_ago', ''),
                    intel['mail'],
                    intel['page'],
                    urlscan_status,
                    intel['ct'],
                    intel['http'],
                    intel['tls'],
                    join_values(perm.get('whois_created')),
                    join_values(perm.get('whois_updated')),
                    join_values(perm.get('whois_expires')),
                    perm.get('whois_registrant', ''),
                    perm.get('whois_org', ''),
                    perm.get('whois_registrar', ''),
                    join_values(perm.get('whois_emails')),
                    perm.get('whois_country', ''),
                    join_values(perm.get('whois_status')),
                    join_values(perm.get('whois_name_servers')),
                    join_values(perm.get('dns_a')),
                    join_values(perm.get('dns_mx')),
                ]

                # WHOIS registrant/org fields are controlled by the very people
                # this tool investigates. Neutralise anything Excel would
                # execute as a formula before it reaches a cell.
                row = [sanitize_spreadsheet_value(value) for value in row]

                ws.append(row)
                
                # Highlight based on risk score
                current_row = ws.max_row
                if risk_score:
                    try:
                        score = int(risk_score)
                        if score >= 70:
                            # High risk - Red
                            risk_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            for cell in ws[current_row]:
                                cell.fill = risk_fill
                                cell.font = Font(color="FFFFFF", bold=True)
                        elif score >= 50:
                            # Medium risk - Orange
                            risk_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                            for cell in ws[current_row]:
                                cell.fill = risk_fill
                        elif score >= 30:
                            # Low-Medium risk - Yellow
                            risk_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            for cell in ws[current_row]:
                                cell.fill = risk_fill
                    except (ValueError, TypeError):
                        pass
                
                # Highlight recent registrations (if not already highlighted by risk)
                elif perm.get('is_recent', False):
                    recent_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    for cell in ws[current_row]:
                        cell.fill = recent_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = 'A2'

    def _create_statistics_sheet(self, wb: Workbook, results: list[dict[str, Any]]) -> None:
        """Create statistics sheet."""
        ws = wb.create_sheet("Statistics")
        
        # Calculate statistics
        total_domains = len(results)
        total_permutations = sum(r['total_permutations'] for r in results)
        total_registered = sum(r['registered_count'] for r in results)
        total_recent = sum(len([p for p in r['permutations'] if p.get('is_recent', False)]) for r in results)
        
        # Count fuzzer types
        fuzzer_counts = {}
        for result in results:
            for perm in result['permutations']:
                fuzzer = perm.get('fuzzer', 'unknown')
                fuzzer_counts[fuzzer] = fuzzer_counts.get(fuzzer, 0) + 1
        
        # Add statistics
        ws.append(['Typo Sniper Scan Statistics'])
        ws.append([])
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        
        ws.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append([])
        
        ws.append(['Overall Statistics'])
        ws['A5'].font = Font(bold=True, size=12)
        ws.append(['Total Domains Scanned:', total_domains])
        ws.append(['Total Permutations Generated:', total_permutations])
        ws.append(['Registered Permutations:', total_registered])
        ws.append(['Recent Registrations:', total_recent])
        ws.append([])
        
        ws.append(['Fuzzer Type Distribution'])
        ws['A11'].font = Font(bold=True, size=12)
        ws.append(['Fuzzer Type', 'Count'])
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        ws['A12'].fill = header_fill
        ws['B12'].fill = header_fill
        ws['A12'].font = header_font
        ws['B12'].font = header_font
        
        for fuzzer, count in sorted(fuzzer_counts.items(), key=lambda x: x[1], reverse=True):
            ws.append([fuzzer, count])
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15


class JSONExporter(BaseExporter):
    """Export results to JSON format."""

    def export(self, results: list[dict[str, Any]], output_dir: Path) -> Path:
        """Export results to JSON file."""
        output_file = self._generate_filename(output_dir, 'json')
        
        # Create export structure
        export_data = {
            'export_date': datetime.now().isoformat(),
            'version': __version__,
            'total_domains': len(results),
            'results': results
        }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        self.logger.info(f"Exported JSON file: {output_file}")
        return output_file


class CSVExporter(BaseExporter):
    """Export results to CSV format."""

    def export(self, results: list[dict[str, Any]], output_dir: Path) -> Path:
        """Export results to CSV file."""
        output_file = self._generate_filename(output_dir, 'csv')
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write headers
            headers = [
                'Scan Date', 'Original Domain', 'Permutation', 'Fuzzer Type',
                'Risk Score', 'Age (days)', 'Mail', 'Page', 'URLScan Status',
                'URLScan Report', 'CT Logs', 'HTTP Status', 'TLS',
                'Created Date', 'Updated Date', 'Expires Date',
                'Registrant', 'Organization', 'Registrar',
                'Emails', 'Country', 'Status',
                'Name Servers', 'IP Addresses', 'Mail Servers', 'Recent'
            ]
            writer.writerow(headers)

            # Write data
            for result in results:
                scan_date = result['scan_date']
                original = result['original_domain']

                for perm in result['permutations']:
                    intel = format_threat_intel(perm)

                    row = [
                        scan_date,
                        original,
                        perm['domain'],
                        perm.get('fuzzer', ''),
                        perm.get('risk_score', ''),
                        perm.get('created_days_ago', ''),
                        intel['mail'],
                        intel['page'],
                        intel['urlscan'],
                        intel['urlscan_url'] or '',
                        intel['ct'],
                        intel['http'],
                        intel['tls'],
                        join_values(perm.get('whois_created')),
                        join_values(perm.get('whois_updated')),
                        join_values(perm.get('whois_expires')),
                        perm.get('whois_registrant', ''),
                        perm.get('whois_org', ''),
                        perm.get('whois_registrar', ''),
                        join_values(perm.get('whois_emails')),
                        perm.get('whois_country', ''),
                        join_values(perm.get('whois_status')),
                        join_values(perm.get('whois_name_servers')),
                        join_values(perm.get('dns_a')),
                        join_values(perm.get('dns_mx')),
                        'Yes' if perm.get('is_recent', False) else 'No',
                    ]

                    # Guard against spreadsheet formula injection via WHOIS data
                    writer.writerow([sanitize_spreadsheet_value(v) for v in row])

        self.logger.info(f"Exported CSV file: {output_file}")
        return output_file


class HTMLExporter(BaseExporter):
    """Export results to HTML format."""

    def export(self, results: list[dict[str, Any]], output_dir: Path) -> Path:
        """Export results to HTML file."""
        output_file = self._generate_filename(output_dir, 'html')
        
        html_content = self._generate_html(results)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        self.logger.info(f"Exported HTML file: {output_file}")
        return output_file

    def _generate_html(self, results: list[dict[str, Any]]) -> str:
        """Generate HTML content."""
        total_registered = sum(r['registered_count'] for r in results)
        total_recent = sum(len([p for p in r['permutations'] if p.get('is_recent', False)]) for r in results)
        
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Typo Sniper Results - {date.today().isoformat()}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: #f5f5f5;
            padding: 20px;
            line-height: 1.6;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #366092;
            margin-bottom: 10px;
            font-size: 2.5em;
        }}
        .subtitle {{
            color: #666;
            margin-bottom: 30px;
            font-size: 1.1em;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 40px;
        }}
        .stat-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-card h3 {{
            font-size: 2em;
            margin-bottom: 5px;
        }}
        .stat-card p {{
            font-size: 0.9em;
            opacity: 0.9;
        }}
        .domain-section {{
            margin-bottom: 40px;
            border: 1px solid #ddd;
            border-radius: 8px;
            overflow: hidden;
        }}
        .domain-header {{
            background: #366092;
            color: white;
            padding: 15px 20px;
            font-size: 1.3em;
            font-weight: bold;
        }}
        .domain-info {{
            background: #f8f9fa;
            padding: 10px 20px;
            display: flex;
            justify-content: space-between;
            border-bottom: 1px solid #ddd;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th {{
            background: #f8f9fa;
            padding: 12px;
            text-align: left;
            font-weight: 600;
            color: #333;
            border-bottom: 2px solid #ddd;
        }}
        td {{
            padding: 10px 12px;
            border-bottom: 1px solid #eee;
        }}
        tr:hover {{
            background: #f8f9fa;
        }}
        .recent {{
            background: #fffde7 !important;
        }}
        .recent::before {{
            content: "🔥 ";
        }}
        .fuzzer-badge {{
            background: #e3f2fd;
            color: #1976d2;
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 0.85em;
            font-weight: 500;
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 2px solid #ddd;
            text-align: center;
            color: #666;
        }}
        code {{
            background: #f5f5f5;
            padding: 2px 6px;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🎯 Typo Sniper Results</h1>
        <p class="subtitle">Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <div class="summary">
            <div class="stat-card">
                <h3>{len(results)}</h3>
                <p>Domains Scanned</p>
            </div>
            <div class="stat-card">
                <h3>{total_registered}</h3>
                <p>Registered Permutations</p>
            </div>
            <div class="stat-card">
                <h3>{total_recent}</h3>
                <p>Recent Registrations</p>
            </div>
        </div>
"""
        
        # Add domain sections
        for result in results:
            recent_count = len([p for p in result['permutations'] if p.get('is_recent', False)])
            
            html_content += f"""
        <div class="domain-section">
            <div class="domain-header">
                {html.escape(str(result['original_domain']))}
            </div>
            <div class="domain-info">
                <span><strong>Scan Date:</strong> {html.escape(str(result['scan_date']))}</span>
                <span><strong>Registered:</strong> {result['registered_count']}</span>
                <span><strong>Recent:</strong> {recent_count}</span>
            </div>
"""
            
            if result['permutations']:
                html_content += """
            <table>
                <thead>
                    <tr>
                        <th>Domain</th>
                        <th>Fuzzer</th>
                        <th>Risk</th>
                        <th>Age (days)</th>
                        <th>Mail</th>
                        <th>Page</th>
                        <th>URLScan Status</th>
                        <th>CT Logs</th>
                        <th>HTTP Status</th>
                        <th>TLS</th>
                        <th>Created</th>
                        <th>Registrant</th>
                        <th>Organization</th>
                        <th>IP</th>
                    </tr>
                </thead>
                <tbody>
"""
                
                for perm in result['permutations']:
                    intel = format_threat_intel(perm)

                    # Everything below originates from third parties: the
                    # domain name itself, WHOIS registrant/org fields, and the
                    # <title> of a page served by the squatter. All of it is
                    # escaped before it reaches the report, so that opening a
                    # report cannot execute attacker-supplied markup.
                    row_class = 'recent' if perm.get('is_recent', False) else ''
                    created = ''
                    if perm.get('whois_created'):
                        created = str(perm['whois_created'][0])

                    ip = ''
                    if perm.get('dns_a'):
                        ip = str(perm['dns_a'][0])

                    if intel['urlscan_url']:
                        urlscan_cell = (
                            f'<a href="{html.escape(intel["urlscan_url"], quote=True)}" '
                            f'target="_blank" rel="noopener noreferrer">'
                            f'{html.escape(intel["urlscan"])}</a>'
                        )
                    else:
                        urlscan_cell = html.escape(intel['urlscan'])

                    age = perm.get('created_days_ago')
                    age_cell = '' if age is None else html.escape(str(age))

                    html_row = f"""
                    <tr class="{html.escape(row_class, quote=True)}">
                        <td><code>{html.escape(str(perm['domain']))}</code></td>
                        <td><span class="fuzzer-badge">{html.escape(str(perm.get('fuzzer', '')))}</span></td>
                        <td>{html.escape(str(perm.get('risk_score', '')))}</td>
                        <td>{age_cell}</td>
                        <td>{html.escape(intel['mail'])}</td>
                        <td>{html.escape(intel['page'])}</td>
                        <td>{urlscan_cell}</td>
                        <td>{html.escape(intel['ct'])}</td>
                        <td>{html.escape(intel['http'])}</td>
                        <td>{html.escape(intel['tls'])}</td>
                        <td>{html.escape(created)}</td>
                        <td>{html.escape(str(perm.get('whois_registrant') or ''))}</td>
                        <td>{html.escape(str(perm.get('whois_org') or ''))}</td>
                        <td>{html.escape(ip)}</td>
                    </tr>
"""
                    html_content += html_row

                html_content += """
                </tbody>
            </table>
"""
            else:
                html_content += """
            <p style="padding: 20px; text-align: center; color: #666;">No registered permutations found</p>
"""
            
            html_content += """
        </div>
"""
        
        html_content += """
        <div class="footer">
            <p>Generated by <strong>Typo Sniper</strong></p>
            <p>Advanced Domain Typosquatting Detection Tool</p>
        </div>
    </div>
</body>
</html>
"""
        
        return html_content
