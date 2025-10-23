"""
Export modules for Typo Sniper results.

Supports multiple output formats: Excel, JSON, CSV, and HTML.
"""

import csv
import json
import logging
from abc import ABC, abstractmethod
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import Config


class BaseExporter(ABC):
    """Base class for all exporters."""

    def __init__(self, config: Config):
        """
        Initialize exporter.

        Args:
            config: Configuration object
        """
        self.config = config
        self.logger = logging.getLogger(self.__class__.__name__)

    @abstractmethod
    def export(self, results: List[Dict[str, Any]], output_dir: Path) -> Path:
        """
        Export results to file.

        Args:
            results: List of scan result dictionaries
            output_dir: Directory to save output file

        Returns:
            Path to the created file
        """
        pass

    def _generate_filename(self, output_dir: Path, extension: str) -> Path:
        """
        Generate output filename with timestamp.

        Args:
            output_dir: Output directory
            extension: File extension (without dot)

        Returns:
            Full path to output file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return output_dir / f"typo_sniper_results_{timestamp}.{extension}"


class ExcelExporter(BaseExporter):
    """Export results to Excel format with rich formatting."""

    def export(self, results: List[Dict[str, Any]], output_dir: Path) -> Path:
        """Export results to Excel file."""
        output_file = self._generate_filename(output_dir, 'xlsx')
        
        wb = Workbook()
        
        # Create summary sheet
        self._create_summary_sheet(wb, results)
        
        # Create detailed results sheet
        self._create_details_sheet(wb, results)
        
        # Create statistics sheet
        self._create_statistics_sheet(wb, results)
        
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(output_file)
        self.logger.info(f"Exported Excel file: {output_file}")
        return output_file

    def _create_summary_sheet(self, wb: Workbook, results: List[Dict[str, Any]]) -> None:
        """Create summary sheet."""
        ws = wb.active
        ws.title = "Summary"
        
        # Headers
        headers = [
            "Scan Date", "Original Domain", "Total Permutations",
            "Registered", "Filtered", "Recent"
        ]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for result in results:
            recent_count = len([p for p in result['permutations'] if p.get('is_recent', False)])
            ws.append([
                result['scan_date'],
                result['original_domain'],
                result['total_permutations'],
                result['registered_count'],
                result['filtered_count'],
                recent_count if recent_count > 0 else ''
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_details_sheet(self, wb: Workbook, results: List[Dict[str, Any]]) -> None:
        """Create detailed results sheet."""
        ws = wb.create_sheet("Details")
        
        # Headers
        headers = [
            "Scan Date", "Original Domain", "Permutation", "Fuzzer Type",
            "Risk Score", "URLScan Status", "CT Logs", "HTTP Status",
            "Created Date", "Updated Date", "Expires Date",
            "Registrant", "Organization", "Registrar",
            "Email", "Country", "Status",
            "Name Servers", "IP Address", "Mail Server"
        ]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for result in results:
            scan_date = result['scan_date']
            original = result['original_domain']
            
            # Add spacing row for each domain
            if ws.max_row > 1:
                ws.append([''] * len(headers))
            
            # Add original domain header
            domain_row = [scan_date, original, '*ORIGINAL*'] + [''] * (len(headers) - 3)
            ws.append(domain_row)
            domain_font = Font(bold=True, color="0066CC", size=11)
            for cell in ws[ws.max_row]:
                cell.font = domain_font
            
            # Add permutations
            for perm in result['permutations']:
                # Get WHOIS data
                created = ', '.join(perm.get('whois_created', []))
                updated = ', '.join(perm.get('whois_updated', []))
                expires = ', '.join(perm.get('whois_expires', []))
                emails = ', '.join(perm.get('whois_emails', []))
                name_servers = ', '.join(perm.get('whois_name_servers', []))
                
                # Get DNS data
                ip = ', '.join(perm.get('dns_a', []))
                mx = ', '.join(perm.get('dns_mx', []))
                
                # Get threat intelligence data
                threat_intel = perm.get('threat_intel', {})
                risk_score = perm.get('risk_score', '')
                urlscan_status = ''
                ct_logs = ''
                http_status = ''
                # Threat intelligence
                if threat_intel:
                    if 'urlscan' in threat_intel:
                        us_data = threat_intel['urlscan']
                        if us_data:
                            # Check for error/status conditions first
                            if 'status' in us_data:
                                status = us_data['status']
                                if status == 'rate_limited':
                                    urlscan_status = "Rate Limited"
                                elif status == 'timeout':
                                    urlscan_status = "Scan Timeout"
                                else:
                                    urlscan_status = f"Error: {status}"
                            else:
                                # URLScan status: Malicious (score) or Clean with link
                                malicious = us_data.get('malicious', False)
                                score = us_data.get('score', 0)
                                report_url = us_data.get('report_url')
                                
                                if malicious or score > 0:
                                    if report_url:
                                        urlscan_status = f"⚠ Malicious ({score}) - {report_url}"
                                    else:
                                        urlscan_status = f"⚠ Malicious ({score})"
                                elif report_url:
                                    urlscan_status = f"✓ Clean ({score}) - {report_url}"
                                else:
                                    # No report_url means scan failed or never existed
                                    urlscan_status = "No Scan Available"
                    
                    if 'certificate_transparency' in threat_intel:
                        ct_data = threat_intel['certificate_transparency']
                        if ct_data:
                            cert_count = ct_data.get('certificates_found', 0)
                            status = ct_data.get('status', '')
                            if cert_count > 0:
                                ct_logs = f"{cert_count} cert(s)"
                            elif status:
                                ct_logs = status
                            else:
                                ct_logs = "0"
                    
                    if 'http_probe' in threat_intel:
                        http_data = threat_intel['http_probe']
                        if http_data:
                            # Get HTTP/HTTPS status
                            http_active = http_data.get('http_active', False)
                            https_active = http_data.get('https_active', False)
                            http_code = http_data.get('http_status', '')
                            https_code = http_data.get('https_status', '')
                            
                            if https_active and https_code:
                                http_status = f"HTTPS: {https_code}"
                            elif http_active and http_code:
                                http_status = f"HTTP: {http_code}"
                            else:
                                http_status = "Inactive"
                
                row = [
                    scan_date,
                    original,
                    perm['domain'],
                    perm.get('fuzzer', ''),
                    risk_score,
                    urlscan_status,
                    ct_logs,
                    http_status,
                    created,
                    updated,
                    expires,
                    perm.get('whois_registrant', ''),
                    perm.get('whois_org', ''),
                    perm.get('whois_registrar', ''),
                    emails,
                    perm.get('whois_country', ''),
                    ', '.join(perm.get('whois_status', [])) if isinstance(perm.get('whois_status'), list) else perm.get('whois_status', ''),
                    name_servers,
                    ip,
                    mx
                ]
                ws.append(row)
                
                # Highlight based on risk score
                current_row = ws.max_row
                if risk_score:
                    try:
                        score = int(risk_score)
                        if score >= 70:
                            # High risk - Red
                            risk_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            for cell in ws[current_row]:
                                cell.fill = risk_fill
                                cell.font = Font(color="FFFFFF", bold=True)
                        elif score >= 50:
                            # Medium risk - Orange
                            risk_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                            for cell in ws[current_row]:
                                cell.fill = risk_fill
                        elif score >= 30:
                            # Low-Medium risk - Yellow
                            risk_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            for cell in ws[current_row]:
                                cell.fill = risk_fill
                    except (ValueError, TypeError):
                        pass
                
                # Highlight recent registrations (if not already highlighted by risk)
                elif perm.get('is_recent', False):
                    recent_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    for cell in ws[current_row]:
                        cell.fill = recent_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = 'A2'

    def _create_statistics_sheet(self, wb: Workbook, results: List[Dict[str, Any]]) -> None:
        """Create statistics sheet."""
        ws = wb.create_sheet("Statistics")
        
        # Calculate statistics
        total_domains = len(results)
        total_permutations = sum(r['total_permutations'] for r in results)
        total_registered = sum(r['registered_count'] for r in results)
        total_recent = sum(len([p for p in r['permutations'] if p.get('is_recent', False)]) for r in results)
        
        # Count fuzzer types
        fuzzer_counts = {}
        for result in results:
            for perm in result['permutations']:
                fuzzer = perm.get('fuzzer', 'unknown')
                fuzzer_counts[fuzzer] = fuzzer_counts.get(fuzzer, 0) + 1
        
        # Add statistics
        ws.append(['Typo Sniper Scan Statistics'])
        ws.append([])
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        
        ws.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append([])
        
        ws.append(['Overall Statistics'])
        ws['A5'].font = Font(bold=True, size=12)
        ws.append(['Total Domains Scanned:', total_domains])
        ws.append(['Total Permutations Generated:', total_permutations])
        ws.append(['Registered Permutations:', total_registered])
        ws.append(['Recent Registrations:', total_recent])
        ws.append([])
        
        ws.append(['Fuzzer Type Distribution'])
        ws['A11'].font = Font(bold=True, size=12)
        ws.append(['Fuzzer Type', 'Count'])
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        ws['A12'].fill = header_fill
        ws['B12'].fill = header_fill
        ws['A12'].font = header_font
        ws['B12'].font = header_font
        
        for fuzzer, count in sorted(fuzzer_counts.items(), key=lambda x: x[1], reverse=True):
            ws.append([fuzzer, count])
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15


class JSONExporter(BaseExporter):
    """Export results to JSON format."""

    def export(self, results: List[Dict[str, Any]], output_dir: Path) -> Path:
        """Export results to JSON file."""
        output_file = self._generate_filename(output_dir, 'json')
        
        # Create export structure
        export_data = {
            'export_date': datetime.now().isoformat(),
            'version': '1.0',
            'total_domains': len(results),
            'results': results
        }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        self.logger.info(f"Exported JSON file: {output_file}")
        return output_file


class CSVExporter(BaseExporter):
    """Export results to CSV format."""

    def export(self, results: List[Dict[str, Any]], output_dir: Path) -> Path:
        """Export results to CSV file."""
        output_file = self._generate_filename(output_dir, 'csv')
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write headers
            headers = [
                'Scan Date', 'Original Domain', 'Permutation', 'Fuzzer Type',
                'Risk Score', 'URLScan Status', 'CT Logs', 'HTTP Status',
                'Created Date', 'Updated Date', 'Expires Date',
                'Registrant', 'Organization', 'Registrar',
                'Emails', 'Country', 'Status',
                'Name Servers', 'IP Addresses', 'Mail Servers', 'Recent'
            ]
            writer.writerow(headers)
            
            # Write data
            for result in results:
                scan_date = result['scan_date']
                original = result['original_domain']
                
                for perm in result['permutations']:
                    # Get threat intelligence data
                    threat_intel = perm.get('threat_intel', {})
                    risk_score = perm.get('risk_score', '')
                    urlscan_status = ''
                    ct_logs = ''
                    http_status = ''
                    
                    if threat_intel:
                        if 'urlscan' in threat_intel:
                            us_data = threat_intel['urlscan']
                            if us_data:
                                # Check for error/status conditions first
                                if 'status' in us_data:
                                    status = us_data['status']
                                    if status == 'rate_limited':
                                        urlscan_status = "Rate Limited"
                                    elif status == 'timeout':
                                        urlscan_status = "Scan Timeout"
                                    else:
                                        urlscan_status = f"Error: {status}"
                                else:
                                    malicious = us_data.get('malicious', False)
                                    score = us_data.get('score', 0)
                                    report_url = us_data.get('report_url')
                                    
                                    if malicious or score > 0:
                                        if report_url:
                                            urlscan_status = f"Malicious ({score}) - {report_url}"
                                        else:
                                            urlscan_status = f"Malicious ({score})"
                                    elif report_url:
                                        urlscan_status = f"Clean ({score}) - {report_url}"
                                    else:
                                        urlscan_status = "No Scan Available"
                        
                        if 'certificate_transparency' in threat_intel:
                            ct_data = threat_intel['certificate_transparency']
                            if ct_data:
                                cert_count = ct_data.get('certificates_found', 0)
                                status = ct_data.get('status', '')
                                if cert_count > 0:
                                    ct_logs = f"{cert_count} cert(s)"
                                elif status:
                                    ct_logs = status
                                else:
                                    ct_logs = "0"
                        
                        if 'http_probe' in threat_intel:
                            http_data = threat_intel['http_probe']
                            if http_data:
                                http_active = http_data.get('http_active', False)
                                https_active = http_data.get('https_active', False)
                                http_code = http_data.get('http_status', '')
                                https_code = http_data.get('https_status', '')
                                
                                if https_active and https_code:
                                    http_status = f"HTTPS: {https_code}"
                                elif http_active and http_code:
                                    http_status = f"HTTP: {http_code}"
                                else:
                                    http_status = "Inactive"
                    
                    row = [
                        scan_date,
                        original,
                        perm['domain'],
                        perm.get('fuzzer', ''),
                        risk_score,
                        urlscan_status,
                        ct_logs,
                        http_status,
                        ', '.join(perm.get('whois_created', [])),
                        ', '.join(perm.get('whois_updated', [])),
                        ', '.join(perm.get('whois_expires', [])),
                        perm.get('whois_registrant', ''),
                        perm.get('whois_org', ''),
                        perm.get('whois_registrar', ''),
                        ', '.join(perm.get('whois_emails', [])),
                        perm.get('whois_country', ''),
                        str(perm.get('whois_status', '')),
                        ', '.join(perm.get('whois_name_servers', [])),
                        ', '.join(perm.get('dns_a', [])),
                        ', '.join(perm.get('dns_mx', [])),
                        'Yes' if perm.get('is_recent', False) else 'No'
                    ]
                    writer.writerow(row)
        
        self.logger.info(f"Exported CSV file: {output_file}")
        return output_file


class HTMLExporter(BaseExporter):
    """Export results to HTML format."""

    def export(self, results: List[Dict[str, Any]], output_dir: Path) -> Path:
        """Export results to HTML file."""
        output_file = self._generate_filename(output_dir, 'html')
        
        html_content = self._generate_html(results)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        self.logger.info(f"Exported HTML file: {output_file}")
        return output_file

    def _generate_html(self, results: List[Dict[str, Any]]) -> str:
        """Generate HTML content."""
        total_registered = sum(r['registered_count'] for r in results)
        total_recent = sum(len([p for p in r['permutations'] if p.get('is_recent', False)]) for r in results)
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Typo Sniper Results - {date.today().isoformat()}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: #f5f5f5;
            padding: 20px;
            line-height: 1.6;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #366092;
            margin-bottom: 10px;
            font-size: 2.5em;
        }}
        .subtitle {{
            color: #666;
            margin-bottom: 30px;
            font-size: 1.1em;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 40px;
        }}
        .stat-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-card h3 {{
            font-size: 2em;
            margin-bottom: 5px;
        }}
        .stat-card p {{
            font-size: 0.9em;
            opacity: 0.9;
        }}
        .domain-section {{
            margin-bottom: 40px;
            border: 1px solid #ddd;
            border-radius: 8px;
            overflow: hidden;
        }}
        .domain-header {{
            background: #366092;
            color: white;
            padding: 15px 20px;
            font-size: 1.3em;
            font-weight: bold;
        }}
        .domain-info {{
            background: #f8f9fa;
            padding: 10px 20px;
            display: flex;
            justify-content: space-between;
            border-bottom: 1px solid #ddd;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th {{
            background: #f8f9fa;
            padding: 12px;
            text-align: left;
            font-weight: 600;
            color: #333;
            border-bottom: 2px solid #ddd;
        }}
        td {{
            padding: 10px 12px;
            border-bottom: 1px solid #eee;
        }}
        tr:hover {{
            background: #f8f9fa;
        }}
        .recent {{
            background: #fffde7 !important;
        }}
        .recent::before {{
            content: "🔥 ";
        }}
        .fuzzer-badge {{
            background: #e3f2fd;
            color: #1976d2;
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 0.85em;
            font-weight: 500;
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 2px solid #ddd;
            text-align: center;
            color: #666;
        }}
        code {{
            background: #f5f5f5;
            padding: 2px 6px;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🎯 Typo Sniper Results</h1>
        <p class="subtitle">Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <div class="summary">
            <div class="stat-card">
                <h3>{len(results)}</h3>
                <p>Domains Scanned</p>
            </div>
            <div class="stat-card">
                <h3>{total_registered}</h3>
                <p>Registered Permutations</p>
            </div>
            <div class="stat-card">
                <h3>{total_recent}</h3>
                <p>Recent Registrations</p>
            </div>
        </div>
"""
        
        # Add domain sections
        for result in results:
            recent_count = len([p for p in result['permutations'] if p.get('is_recent', False)])
            
            html += f"""
        <div class="domain-section">
            <div class="domain-header">
                {result['original_domain']}
            </div>
            <div class="domain-info">
                <span><strong>Scan Date:</strong> {result['scan_date']}</span>
                <span><strong>Registered:</strong> {result['registered_count']}</span>
                <span><strong>Recent:</strong> {recent_count}</span>
            </div>
"""
            
            if result['permutations']:
                html += """
            <table>
                <thead>
                    <tr>
                        <th>Domain</th>
                        <th>Fuzzer</th>
                        <th>Risk</th>
                        <th>URLScan Status</th>
                        <th>CT Logs</th>
                        <th>HTTP Status</th>
                        <th>Created</th>
                        <th>Registrant</th>
                        <th>IP</th>
                    </tr>
                </thead>
                <tbody>
"""
                
                for perm in result['permutations']:
                    row_class = 'recent' if perm.get('is_recent', False) else ''
                    created = perm.get('whois_created', [''])[0] if perm.get('whois_created') else ''
                    registrant = perm.get('whois_registrant', '')
                    ip = perm.get('dns_a', [''])[0] if perm.get('dns_a') else ''
                    
                    # Get threat intelligence data
                    threat_intel = perm.get('threat_intel', {})
                    risk_score = perm.get('risk_score', '')
                    urlscan_status = ''
                    ct_logs = ''
                    http_status = ''
                    
                    if threat_intel:
                        if 'urlscan' in threat_intel:
                            us_data = threat_intel['urlscan']
                            if us_data:
                                # Check for error/status conditions first
                                if 'status' in us_data:
                                    status = us_data['status']
                                    if status == 'rate_limited':
                                        urlscan_status = "⏱️ Rate Limited"
                                    elif status == 'timeout':
                                        urlscan_status = "⏱️ Scan Timeout"
                                    else:
                                        urlscan_status = f"❌ Error: {status}"
                                else:
                                    malicious = us_data.get('malicious', False)
                                    score = us_data.get('score', 0)
                                    report_url = us_data.get('report_url')
                                    
                                    if malicious or score > 0:
                                        if report_url:
                                            urlscan_status = f"⚠️ <a href='{report_url}' target='_blank'>Malicious ({score})</a>"
                                        else:
                                            urlscan_status = f"⚠️ Malicious ({score})"
                                    elif report_url:
                                        urlscan_status = f"✓ <a href='{report_url}' target='_blank'>Clean ({score})</a>"
                                    else:
                                        urlscan_status = "⚪ No Scan Available"
                        
                        if 'certificate_transparency' in threat_intel:
                            ct_data = threat_intel['certificate_transparency']
                            if ct_data:
                                cert_count = ct_data.get('certificates_found', 0)
                                status = ct_data.get('status', '')
                                if cert_count > 0:
                                    ct_logs = f"✓ {cert_count} cert(s)"
                                elif status:
                                    ct_logs = status
                                else:
                                    ct_logs = "0"
                        
                        if 'http_probe' in threat_intel:
                            http_data = threat_intel['http_probe']
                            if http_data:
                                http_active = http_data.get('http_active', False)
                                https_active = http_data.get('https_active', False)
                                http_code = http_data.get('http_status', '')
                                https_code = http_data.get('https_status', '')
                                
                                if https_active and https_code:
                                    http_status = f"🔒 HTTPS: {https_code}"
                                elif http_active and http_code:
                                    http_status = f"HTTP: {http_code}"
                                else:
                                    http_status = "Inactive"
                    
                    html += f"""
                    <tr class="{row_class}">
                        <td><code>{perm['domain']}</code></td>
                        <td><span class="fuzzer-badge">{perm.get('fuzzer', '')}</span></td>
                        <td>{risk_score}</td>
                        <td>{urlscan_status}</td>
                        <td>{ct_logs}</td>
                        <td>{http_status}</td>
                        <td>{created}</td>
                        <td>{registrant}</td>
                        <td>{ip}</td>
                    </tr>
"""
                
                html += """
                </tbody>
            </table>
"""
            else:
                html += """
            <p style="padding: 20px; text-align: center; color: #666;">No registered permutations found</p>
"""
            
            html += """
        </div>
"""
        
        html += """
        <div class="footer">
            <p>Generated by <strong>Typo Sniper v1.0</strong></p>
            <p>Advanced Domain Typosquatting Detection Tool</p>
        </div>
    </div>
</body>
</html>
"""
        
        return html
